# 读入数据
import sys
import pandas as pd
import openpyxl
import numpy as np
import requests
import json
import threading

from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5 import QtWidgets, QtGui
from PyQt5.QtWidgets import *

import ui
from ui import Ui_MainWindow

class Mywindow(QtWidgets.QMainWindow, ui.Ui_MainWindow):

    def __init__(self):
        super(Mywindow, self).__init__()
        # QtWidgets.QMainWindow.__init__(self)
        self.setupUi(self)

        self.timer = threading.Timer(0.5, self.fun_timer)
        self.timer.start()

        pd.set_option('display.unicode.ambiguous_as_wide', True)
        pd.set_option('display.unicode.east_asian_width', True)

        self.book = openpyxl.load_workbook('./1.xlsx')
        self.sheet = self.book.worksheets[0]
        self.currRow = 0
        self.newRow = 0

        pd.set_option('max_colwidth', 20)  # 设置value的显示长度为100，默认为50

        c = ""
        i = 1
        while(True):
            print ("row :" ,i)
            print("row[i].value: ", self.sheet.cell(row = i, column = 2).value)
            print("row[i].value.bool: ", self.sheet.cell(row = i, column = 3).value)
            
            if(None == self.sheet.cell(row = i, column = 1).value):
                c = self.retC(i)
                self.currRow = i
                self.newRow = i
                break
            i += 1
        print("self.currRow:",self.currRow, " self.newRow:",self.newRow)
        self.textBrowser.append(str(c))
        self.textBrowser.moveCursor(self.textBrowser.textCursor().End)

    def retC(self, i):
        c = "<font size=\"6\">laser_dis : " + \
            str(self.sheet.cell(row = i, column = 1).value) + '<br />' + \
            "aim_tract : " + \
            str(self.sheet.cell(row = i, column = 2).value) + '<br />' + \
            "</font>"
        if(None != self.sheet.cell(row = i, column = 3).value):
            c += "<br /><br /><font size=\"6\" color=\"red\" > label :   "
            c += str(self.sheet.cell(row = i, column = 3).value)
            c += "</font>"
        return c

    def showCurr(self):
        self.textBrowser.clear()
        c = self.retC(self.currRow)
        self.textBrowser.append(str(c))
        self.textBrowser.moveCursor(self.textBrowser.textCursor().End)

    def showNext(self):
        self.textBrowser.clear()
        self.currRow += 1
        c = self.retC(self.currRow)
        self.textBrowser.append(str(c))
        self.textBrowser.moveCursor(self.textBrowser.textCursor().End)

    def saveExcel(self):
        self.book.save('./1.xlsx')

    def p1_ck(self):  # 上一个
        self.textBrowser.clear()
        self.currRow -= 1
        if(self.currRow < 0):
            self.currRow = 0
        c = self.retC(self.currRow)
        self.textBrowser.append(str(c))
        self.textBrowser.moveCursor(self.textBrowser.textCursor().End)

    def p2_ck(self):  # 下一个
        self.showNext()
    
    def p3_ck(self):  # P
        self.sheet.cell(row = self.currRow, column = 3).value = 1
        self.saveExcel()
        self.showCurr()
    
    def p4_ck(self):  # N
        self.sheet.cell(row = self.currRow, column = 3).value = 0
        self.saveExcel()
        self.showCurr()

    def fun_timer(self):
        response = requests.get("http://192.168.31.80/excel.json")
        data2 = json.loads(response.text)
        
        print(data2)
        if(data2['data']['laser_dis'] != 0 and data2['data']['aim_tract'] != 0): #new data coming
            print("Hello", self.newRow -1)
            if(self.sheet.cell(row = self.newRow -1, column = 3).value == None): #上一个数据如果没标，自动给0
                self.sheet.cell(row = self.newRow -1, column = 3).value = 0
            else:
                print(data2['data']['laser_dis'], data2['data']['aim_tract'])
                self.sheet.cell(row = self.newRow, column = 1).value = data2['data']['laser_dis']
                self.sheet.cell(row = self.newRow, column = 2).value = data2['data']['aim_tract']
                self.newRow += 1
                self.currRow = self.newRow
                self.showCurr()
        timer = threading.Timer(0.5, self.fun_timer)
        timer.start()

# MAIN
app = QtWidgets.QApplication(sys.argv)
window = Mywindow()
window.show()
sys.exit(app.exec_())